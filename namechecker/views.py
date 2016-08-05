import openpyxl
from django.views.generic import View
from django.template.loader import render_to_string
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
import os
from .forms import ImportFile
import tempfile
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

class Checker:

    def setdata(self,file,sheet_name,row_name,row_num):
        """ 엑셀의 속성값으로 구조 파악"""

        def findtarget(head_col,row_name):
            """ 식물 이름 목록이 있는 열 위치값을 반환 """
            for col in head_col:
                if col.value == row_name:
                    return {'index':col.col_idx , 'str':col.column }

        self.file = openpyxl.load_workbook(file)
        self.sheet = self.file.get_sheet_by_name(sheet_name)
        self.row_name = row_name
        self.head_col = self.sheet.rows[row_num]
        self.head_row = findtarget(self.head_col,row_name)
        self.count_list = self.sheet.max_row

        # 매칭된 이름을 저장할 셀의 위치
        self._indx = self.sheet.max_column
        self.row_nameko = get_column_letter(self._indx+1)
        self.row_nameen = get_column_letter(self._indx+2)

    def search(self):
        for i in range(1,self.count_list):
            name = self.sheet.cell(row=i,column=self.head_row['index']).value
            self.sheet['{0}{1}'.format(self.row_nameko,i)] = name
            self.sheet['{0}{1}'.format(self.row_nameen,i)] = name



def checkname(file):
    """
    이름을 매칭하여 알려준다
    :param file:
    :return:
    """
    pass


class Index(View):
    def get(self, request, data=None):
        form = ImportFile()
        context = {'test':'filecheck','upload':'no', 'form':form}

        data = render_to_string('check/index.html',context,request=request)
        return HttpResponse(data)

    def post(self, request):
        form = ImportFile(request.POST,request.FILES)
        test = 'filecheck'
        if form.is_valid(): # 파일이 업로드 되었을 경우
            # 업로드한 파일및 엑셀 속성값 호출
            file = request.FILES['file']    # 업로드된 파일
            sheet_name = request.POST['sheetname'] # 식물 목록이 존해하는 시트이름
            row_name = request.POST['rowname'] # 식물 목록의 머릿행 이름
            # 머릿행의 번호
            if request.POST['rownum']:
               row_num = int(request.POST['rownum'])-1
            else:
                row_num = 0

            # 파일 형식확인 : 엑셀파일이 아닐경우
            if file.content_type != 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
                upload = 'no'
                test = file.content_type
                message = '엑셀파일이 아닌거 같습니다. 파일 확장자를 확인해주시고 업로드 하세요'
            else:
                # 업로드 파일을 임시파일로 저장하기
                # tmp = tempfile.NamedTemporaryFile()
                # tmp.write(file.read())
                # 엑셀 파일을 불러 시트 목록 출력하기
                try:
                    target = Checker()
                    target.setdata(file,sheet_name,row_name,row_num)
                    target.search()
                    target.file.save('test.xlsx')
                    upload = 'yes'
                    message = '파일이 성공적으로 업로드 되었습니다. 잠시만 기다리시면 작업이 완료된 파일을 다운 받을수 있습니다.'
                    # test code
                    message = '{0},{1},{2}'.format(target.sheet,target.head_col,target.head_row)
                except:
                    upload = 'no'
                    message = '지원하지 않는 엑셀 구조 입니다.관리자 에게 문의해 주세요'


        else:   # 업로드 파일이 없을 경우
            upload = 'no'
            message = '이런, 파일을 올리신게 맞나요? 저는 아무것도 받지 못했습니다.. 다시 확인해 주세요'
        form = ImportFile()
        context = {'test': test, 'upload': upload, 'message': message, 'form':form}
        data = render_to_string('check/index.html', context, request=request)
        return HttpResponse(data)