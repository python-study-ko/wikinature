import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from .plantsnameAPI import matchlist

class Checker:

    def setdata(self,file,sheet_name,row_name,row_num):
        """ 엑셀의 속성값으로 구조 파악"""

        def findtarget(head_col,row_name):
            """ 식물 이름 목록이 있는 열 위치값을 반환 """
            for col in head_col:
                if col.value == row_name:
                    return {'index':col.col_idx , 'str':col.column }

        self.file = openpyxl.load_workbook(file)
        self.row_num = row_num
        self.row_name = row_name
        try:
            self.sheet = self.file.get_sheet_by_name(sheet_name)
        except KeyError as e:
            raise e
        self.head_col = self.sheet.rows[row_num]
        self.head_row = findtarget(self.head_col,row_name)
        self.count_list = self.sheet.max_row

        # 매칭된 이름을 저장할 셀의 위치
        self._indx = self.sheet.max_column
        self.row_nameko = get_column_letter(self._indx+1)
        self.row_nameen = get_column_letter(self._indx+2)

    def search(self):
        # 중복되지 않는 식물명 리스트를 만든
        name_list = []
        for i in range(self.row_num+2,self.count_list+1):
            name = self.sheet.cell(row=i,column=self.head_row['index']).value
            if name not in name_list:
                name_list.append(name)

        # 식물이름에 매칭되는 국,학명 정보를 가져온다
        plant_info = matchlist(name_list)

        # 국,학명 머리열 이름 추가
        self.sheet['{0}{1}'.format(self.row_nameko, self.row_num + 1)] = '국명'
        self.sheet['{0}{1}'.format(self.row_nameen, self.row_num + 1)] = '학명'

        for i in range(self.row_num+2,self.count_list+1):
            name = self.sheet.cell(row=i,column=self.head_row['index']).value
            if plant_info[name]['state'] == False:
                self.sheet['{0}{1}'.format(self.row_nameko, i)] = plant_info[name]['etc']
            else:
                self.sheet['{0}{1}'.format(self.row_nameko,i)] = plant_info[name]['국명']
                self.sheet['{0}{1}'.format(self.row_nameen,i)] = plant_info[name]['학명']